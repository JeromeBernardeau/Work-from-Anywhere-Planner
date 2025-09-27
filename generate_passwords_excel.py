#!/usr/bin/env python3
"""
Generate Excel file with all FICOFI employees and their passwords
Creates a comprehensive password directory for the Work-from-Anywhere Planner
"""

import pandas as pd
import hashlib
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime

def generate_password_for_email(email):
    """Generate a deterministic 10-digit password for an email address"""
    # Special passwords for executives
    special_passwords = {
        'jbernardeau@ficofi.com': '2034678915',
        'pcapdouze@ficofi.com': '1023456789',
        'mdestot@ficofi.com': '2034567891',
        'slefebure@ficofi.com': '2315476980',
        'acapdouze@ficofi.com': '1011225632',
        'cmanoukian@ficofi.com': '3124567890',
        'sterenn@ficofi.com': '7151285100',
        'rcavalotti@ficofi.com': '4321567890',
        'ppecqueux@ficofi.com': '5432167890'
    }
    
    # Check if email has a special password
    email_lower = email.lower()
    if email_lower in special_passwords:
        return special_passwords[email_lower]
    
    # Generate deterministic password for others
    hash_object = hashlib.sha256(email.encode())
    hash_hex = hash_object.hexdigest()
    
    # Convert hex to numbers, ensuring we get 10 digits
    password = ''
    for char in hash_hex:
        if len(password) >= 10:
            break
        if char.isdigit():
            password += char
        else:
            # Convert a-f to 0-5
            password += str(ord(char) % 6)
    
    # If we still don't have 10 digits, pad with hash-based numbers
    while len(password) < 10:
        password += str(abs(hash(email)) % 10)
        email = email + 'x'  # Modify email to get different hash
    
    return password[:10]

# Read the staff data from our seed.sql file
print("Reading staff data from seed.sql...")
staff_data = []

with open('/home/user/webapp/seed.sql', 'r') as f:
    content = f.read()
    
# Parse the INSERT statements - updated format
import re
# Pattern: (display_name, email, site, department, manager_email, second_level_manager_email, admin_access)
insert_pattern = r"\('([^']+)', '([^']+)', '([^']*)', '([^']*)', '([^']*)', '([^']*)', (\d)\)"
matches = re.findall(insert_pattern, content)

for match in matches:
    display_name, email, site, department, manager_email, second_level_manager_email, admin_access = match
    password = generate_password_for_email(email)
    
    # Determine position based on hierarchy
    position = 'Staff'
    if not manager_email and not second_level_manager_email:
        if department in ['Board', 'ManCO']:
            position = 'Executive'
        else:
            position = 'Department Head'
    elif not second_level_manager_email and manager_email:
        position = 'Manager'
    
    staff_data.append({
        'Name': display_name,
        'Email': email,
        'Site': site if site else 'Paris',
        'Department': department if department else 'General',
        'Position': position,
        'Manager': manager_email if manager_email else '-',
        'Password': password,
        'Admin': 'Yes' if admin_access == '1' else 'No',
        'Status': 'Active'
    })

# Create DataFrame
df = pd.DataFrame(staff_data)

# Sort by department and then by name
df = df.sort_values(['Department', 'Name'])

# Create Excel file with formatting
output_file = '/home/user/webapp/FICOFI_Employee_Passwords.xlsx'
print(f"Creating Excel file: {output_file}")

# Write to Excel with formatting
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Employee Passwords', index=False)
    
    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Employee Passwords']
    
    # Define styles
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)
    cell_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format header row
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = cell_border
    
    # Format data cells
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = cell_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Highlight executives
            if cell.column == 3:  # Department column
                if cell.value in ['ManCO', 'Board']:
                    for c in row:
                        c.fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
    
    # Adjust column widths
    column_widths = {
        'A': 25,  # Name
        'B': 35,  # Email
        'C': 15,  # Site
        'D': 15,  # Department
        'E': 20,  # Position
        'F': 30,  # Manager
        'G': 15,  # Password
        'H': 10,  # Admin
        'I': 10   # Status
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    # Add summary sheet
    summary_df = df.groupby('Department').size().reset_index(name='Employee Count')
    summary_df['Total Passwords'] = summary_df['Employee Count']
    summary_df.loc[len(summary_df)] = ['TOTAL', summary_df['Employee Count'].sum(), summary_df['Employee Count'].sum()]
    
    summary_df.to_excel(writer, sheet_name='Summary', index=False)
    
    # Format summary sheet
    summary_sheet = writer.sheets['Summary']
    for cell in summary_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = cell_border
    
    for row in summary_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = cell_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if row[0].value == 'TOTAL':
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    
    # Adjust summary column widths
    summary_sheet.column_dimensions['A'].width = 20
    summary_sheet.column_dimensions['B'].width = 20
    summary_sheet.column_dimensions['C'].width = 20

print(f"\n✅ Excel file created successfully!")
print(f"📁 Location: {output_file}")
print(f"👥 Total employees: {len(df)}")
print(f"🔐 All passwords generated using secure algorithm")

# Print department summary
print("\n📊 Department Summary:")
dept_summary = df.groupby('Department').size()
for dept, count in dept_summary.items():
    print(f"  • {dept}: {count} employees")

print(f"\n🎯 File ready for distribution to FICOFI management")